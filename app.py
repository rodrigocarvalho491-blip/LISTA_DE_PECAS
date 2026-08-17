import streamlit as st
import openpyxl
from openpyxl.styles import Font
import pandas as pd
import io
import datetime
import re
import os
from streamlit_mic_recorder import mic_recorder
import speech_recognition as sr

# 1. Configuração da página
st.set_page_config(page_title="Gerador de Lista de Peças", page_icon="📦", layout="wide")

# 2. Identidade Visual Customizada (CSS)
CUSTOM_CSS = """
<style>
    h1, h2, h3 {
        color: #004080 !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    div.stButton > button:first-child {
        background-color: #004080 !important;
        color: #ffffff !important;
        border-radius: 8px !important;
        border: none !important;
        padding: 10px 24px !important;
        font-weight: bold !important;
        font-size: 16px !important;
        transition: all 0.3s ease;
    }
    div.stButton > button:first-child:hover {
        background-color: #002b55 !important;
        color: #ffffff !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    .stTextInput > div > div > input, .stTextArea textarea {
        border-radius: 6px !important;
        border: 1px solid #cccccc !important;
    }
    .stTextInput > div > div > input:focus, .stTextArea textarea:focus {
        border-color: #004080 !important;
    }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# 3. Logo e Cabeçalho
LOGO_PATH = "logo.png"

col_logo, col_titulo = st.columns([1, 4])
with col_logo:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=160)
    else:
        st.caption("📷 *Adicione logo.png na pasta*")

with col_titulo:
    st.title("Gerador de Lista de Peças")
    st.markdown("Sistema interno de preenchimento automatizado de ordens de peças.")

st.divider()

# --- GUIA RÁPIDO DE EQUIVALÊNCIAS DE MEDIDAS ---
with st.expander("💡 **Guia Rápido de Equivalência de Medidas (Multicamada x Aço x Cobre)**"):
    st.markdown("""
    | Tubo Multicamada | Tubo de Aço (Rosca BSP) | Tubo de Cobre |
    | :---: | :---: | :---: |
    | **16 mm** | 3/8" | 15 mm |
    | **20 mm** | 1/2" | 22 mm |
    | **26 mm** | 3/4" | 28 mm |
    | **32 mm** | 1" | 35 mm |
    | **40 mm** | 1.1/4" | 42 mm |
    | **50 mm** | 1.5" | 54 mm |
    | **63 mm** | 2" | 66 mm |
    """)

st.divider()

# --- DADOS DO CLIENTE ---
st.subheader("1. Dados do Cliente (Opcional)")
col1, col2, col3 = st.columns(3)

with col1:
    nome_cliente = st.text_input("Nome do Cliente", placeholder="Ex: João da Silva")
with col2:
    cod_cliente = st.text_input("Código do Cliente", placeholder="Ex: 1042")
with col3:
    num_os = st.text_input("Nº da OS", placeholder="Ex: OS-2026/08")

endereco = st.text_input("Endereço", placeholder="Ex: Av. Principal, 500 - SJC")
data_hoje = st.date_input("Data", datetime.date.today()).strftime("%d/%m/%Y")

st.divider()

# --- ENTRADA DA LISTA DE PEÇAS ---
st.subheader("2. Lista de Peças")

col_mic, col_info = st.columns([1, 4])

with col_mic:
    audio = mic_recorder(
        start_prompt="🎙️ Gravar Voz",
        stop_prompt="⏹️ Finalizar",
        key='recorder'
    )

texto_ditado = ""

if audio and 'bytes' in audio:
    r = sr.Recognizer()
    audio_file = io.BytesIO(audio['bytes'])
    with sr.AudioFile(audio_file) as source:
        audio_data = r.record(source)
        try:
            texto_ditado = r.recognize_google(audio_data, language="pt-BR")
            st.success(f"🗣️ Reconhecido: **{texto_ditado}**")
        except Exception:
            st.error("❌ Não foi possível reconhecer o áudio. Tente falar novamente.")

texto_solicitacao = st.text_area(
    "Digite, cole ou dite os itens aqui:",
    value=texto_ditado if texto_ditado else "",
    height=150,
    placeholder="Ex: 25m tubo multi 20mm, 25 suportes tubo multi 20mm, 02 niple 3/4\" x 1/2\", 1 cotovelo 1 1/2\"",
)

formato = st.radio(
    "Escolha o formato de saída:",
    ["📄 Texto para Mensagem / WhatsApp", "📊 Planilha Excel (Modelo Base)"],
    horizontal=True
)

st.divider()

def normalize_text(text):
    text = text.lower()
    text = re.sub(r'[áàãâä]', 'a', text)
    text = re.sub(r'[éèêë]', 'e', text)
    text = re.sub(r'[íìîï]', 'i', text)
    text = re.sub(r'[óòõôö]', 'o', text)
    text = re.sub(r'[úùûü]', 'u', text)
    text = re.sub(r'[ç]', 'c', text)
    text = re.sub(r'[^a-z0-9\s/"]', ' ', text)
    return " ".join(text.split())

def match_catalog_item(user_item_str, catalog):
    norm_user = normalize_text(user_item_str)
    
    scores = []
    for cat in catalog:
        code = cat['code']
        desc = cat['desc']
        norm_desc = normalize_text(desc)
        score = 0
        
        # --- NIPLES ---
        if "niple" in norm_user:
            if "3/4" in norm_user and "1/2" in norm_user and code == "PI1370":
                score += 300
            elif "1/2" in norm_user and "3/4" not in norm_user and code == "PI1230":
                score += 200
            elif "3/4" in norm_user and "1/2" not in norm_user and code == "PI1240":
                score += 200

        # --- SUPORTES ---
        elif "suporte" in norm_user:
            if "20" in norm_user and code == "PI3569":
                score += 200
            elif "26" in norm_user and code == "PI3571":
                score += 200
            elif ("l" in norm_user or "15" in norm_user) and code == "PI5510":
                score += 200

        # --- CONECTORES ---
        elif "conector" in norm_user:
            if "macho" in norm_user and "20" in norm_user and code == "PI3651":
                score += 200
            elif "femea" in norm_user and "20" in norm_user and code == "PI3644":
                score += 200

        # --- COTOVELOS ---
        elif "cotovelo" in norm_user:
            if "1 1/2" in norm_user and code == "PI3001":
                score += 300
            elif "femea" in norm_user and code == "PI3574":
                score += 200
            elif "20" in norm_user and code == "PI3545":
                score += 200
            elif "26" in norm_user and code == "PI3546":
                score += 200
            elif ("1/2" in norm_user or "meia" in norm_user) and "1 1/2" not in norm_user and code == "PI0600":
                score += 200

        # --- TUBOS ---
        elif "tubo" in norm_user:
            if "multi" in norm_user and "20" in norm_user and code == "PI3668":
                score += 200
            elif "multi" in norm_user and "26" in norm_user and code == "PI3669":
                score += 200
            elif "aco" in norm_user and "1 1/2" in norm_user and code == "PI0050":
                score += 200
            elif "aco" in norm_user and code == "PI0010":
                score += 200

        # --- TÊS ---
        elif "te" in norm_user or " t " in f" {norm_user} ":
            if "20" in norm_user and code == "PI3549":
                score += 200
            elif "26" in norm_user and code == "PI3551":
                score += 200
            elif code == "PI0990":
                score += 150

        # --- LUVAS ---
        elif "luva" in norm_user:
            if "1 1/2" in norm_user and code == "PI1461":
                score += 300
            elif "26" in norm_user and code == "PI3593":
                score += 200
            elif "20" in norm_user and code == "PI3592":
                score += 200

        # --- OUTRAS PEÇAS ---
        elif ("valvula" in norm_user or "esferica" in norm_user) and code == "PI2190":
            score += 200
        elif ("caps" in norm_user or "tampao" in norm_user) and code == "PI2970":
            score += 200
        elif "uniao" in norm_user and code == "PI6673":
            score += 200

        user_words = set(norm_user.split())
        desc_words = set(norm_desc.split())
        common = user_words.intersection(desc_words)
        
        for w in user_words:
            if w in desc_words:
                score += 5
                if "/" in w or "mm" in w or w.isdigit():
                    score += 25

        scores.append((score, cat))
        
    scores.sort(key=lambda x: x[0], reverse=True)
    if scores and scores[0][0] > 0:
        return scores[0][1]
    return None

def process_items(texto_solicitacao, catalog):
    lines = [l.strip() for l in re.split(r'[\n,]', texto_solicitacao) if l.strip()]
    results = []
    
    for line in lines:
        m = re.match(r'^(\d+(?:[\.,]\d+)?)\s*(m|metros|un|unidades|unidade|pc|pcs)?\s*(.*)$', line, re.IGNORECASE)
        if m:
            qty_num = m.group(1)
            unit = m.group(2) or ""
            item_str = m.group(3).strip()
            
            try:
                clean_val = str(int(float(qty_num.replace(',', '.'))))
            except:
                clean_val = qty_num
                
            if unit.lower() in ['m', 'metros']:
                qty_formatted = f"{clean_val}m"
            else:
                qty_formatted = clean_val
        else:
            qty_formatted = "1"
            clean_val = "1"
            item_str = line
            
        cat_match = match_catalog_item(item_str, catalog)
        if cat_match:
            results.append({
                "qty_str": qty_formatted,
                "qty_num": clean_val,
                "code": cat_match['code'],
                "desc": cat_match['desc'],
                "row": cat_match.get('row'),
                "col_qty": cat_match.get('col_qty', 1)
            })
        else:
            results.append({
                "qty_str": qty_formatted,
                "qty_num": clean_val,
                "code": "OUTROS",
                "desc": item_str,
                "row": None,
                "col_qty": 1
            })
    return results

if st.button("🚀 Processar Solicitação"):
    if not texto_solicitacao.strip():
        st.warning("⚠️ Por favor, informe ao menos um item na lista.")
    else:
        catalog = []
        try:
            wb = openpyxl.load_workbook("lista_peças.xlsx")
            ws = wb.active
            for r in range(1, ws.max_row + 1):
                c_code = ws.cell(row=r, column=2).value
                c_desc = ws.cell(row=r, column=3).value
                if c_code and str(c_code).strip().startswith('PI'):
                    catalog.append({"row": r, "col_qty": 1, "code": str(c_code).strip(), "desc": str(c_desc).strip()})
                
                c_code2 = ws.cell(row=r, column=6).value
                c_desc2 = ws.cell(row=r, column=7).value
                if c_code2 and str(c_code2).strip().startswith('PI'):
                    catalog.append({"row": r, "col_qty": 5, "code": str(c_code2).strip(), "desc": str(c_desc2).strip()})
        except FileNotFoundError:
            st.error("❌ Arquivo 'lista_peças.xlsx' não encontrado no diretório do projeto.")
            st.stop()
            
        items_processados = process_items(texto_solicitacao, catalog)
        
        if "Texto" in formato:
            st.subheader("📋 Lista em Texto Formatada")
            linhas_resultado = [f"{item['qty_str']} - {item['code']} - {item['desc']}" for item in items_processados]
            texto_final = "\n".join(linhas_resultado)
            st.text_area("Resultado (Copie e cole):", value=texto_final, height=250)
            
        else:
            try:
                wb = openpyxl.load_workbook("lista_peças.xlsx")
                ws = wb.active
                
                fonte_cabecalho = Font(name="Calibri", size=18, bold=True)
                
                if data_hoje:
                    ws["A4"] = f" Data: {data_hoje}"
                    ws["A4"].font = fonte_cabecalho
                if nome_cliente.strip():
                    ws["C4"] = f"Cliente: {nome_cliente}"
                    ws["C4"].font = fonte_cabecalho
                if num_os.strip():
                    ws["E4"] = f"OS: {num_os}"
                    ws["E4"].font = fonte_cabecalho
                if cod_cliente.strip():
                    ws["G4"] = f"Cód. Cliente: {cod_cliente}"
                    ws["G4"].font = fonte_cabecalho
                if endereco.strip():
                    ws["A5"] = f" Endereço: {endereco}"
                    ws["A5"].font = fonte_cabecalho

                for item in items_processados:
                    if item['row'] and item['col_qty']:
                        ws.cell(row=item['row'], column=item['col_qty'], value=item['qty_num'])

                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                st.success("✅ Planilha gerada com sucesso!")
                nome_arq = f"Lista_{nome_cliente.replace(' ', '_')}.xlsx" if nome_cliente.strip() else "Lista_de_Pecas_Preenchida.xlsx"
                
                st.download_button(
                    label="📥 Baixar Planilha Excel (.xlsx)",
                    data=buffer,
                    file_name=nome_arq,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ Erro ao preencher a planilha: {e}")
